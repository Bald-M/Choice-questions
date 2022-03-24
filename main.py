# -*- coding:utf-8 -*-
# 作者：张子涵
# 时间：2022/3/24

from openpyxl import load_workbook
from random import randint

# 读取Excel的数据
# 1,1表示A列 2,2表示B列 3,3表示C列
def read_data(min_col,max_col):
    data_list = []
    for datas in ws.iter_cols(min_col=min_col, max_col=max_col):
        for data in datas:
            data_list.append(data.value)
        return data_list

#
def ask_question(rand):
    # 从questions列表中随机抽取问题并赋值给question
    question = questions[rand]
    # 从options列表中随机抽取问题并赋值给option
    option = options[rand]
    # 打印问题和选项
    print(question)
    print(option)

if __name__ == '__main__':
    # 读取工作簿
    wb = load_workbook("题库.xlsx")
    # 激活工作表
    ws = wb.active

    questions = read_data(1, 1)
    options = read_data(2, 2)
    answers = read_data(3, 3)

    rand = randint(1, int(ws.max_row) - 1)
    while True:
        answer = answers[rand]
        ask_question(rand)
        answer_question = input("请输入答案：")
        if answer_question.lower() == answer.lower():
            print("您的答案正确！")
            rand = randint(1, int(ws.max_row) - 1)
        elif len(answer_question.lower()) < len(answer.lower()) or len(answer_question.lower()) < len(answer.lower()):
            print("本题为多选题，有%s个答案，您选择了%s个答案"%(len(answer.lower()),len(answer_question.lower())))
            print("请重新选择！")
        elif answer_question.lower() != answer.lower():
            print("您的答案错误！")
            print("请重新选择！")
        else:
            print("请输入正确的字符")









